from __future__ import annotations

import calendar
from datetime import date, timedelta

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.core.exceptions import NotFoundError, ValidationError
from app.core.logger import get_logger
from app.repositories.daily_attendance_repo import DailyAttendanceRepository
from app.repositories.employee_repo import EmployeeRepository
from app.utils.excel_utils import format_seconds, workbook_to_bytes, write_sheet
from app.utils.time_utils import to_local

log = get_logger(__name__)

_MAX_RANGE_DAYS = 366


class ReportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.daily_repo = DailyAttendanceRepository(db)
        self.employee_repo = EmployeeRepository(db)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _fmt(dt) -> str:
        if dt is None:
            return ""
        return to_local(dt).strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def _assert_range(start: date, end: date) -> None:
        if end < start:
            raise ValidationError("end_date must be >= start_date")
        if (end - start).days > _MAX_RANGE_DAYS:
            raise ValidationError(f"Date range exceeds {_MAX_RANGE_DAYS} days — narrow the request")

    # ------------------------------------------------------------------
    # Daily — one row per employee who had activity
    # ------------------------------------------------------------------

    def daily_report(self, work_date: date) -> bytes:
        sessions = self.daily_repo.list_by_date(work_date)
        employees = self.employee_repo.get_by_ids(list({s.employee_id for s in sessions}))

        headers = [
            "Employee Code",
            "Name",
            "Department",
            "Status",
            "IN",
            "First Break Out",
            "First Break In",
            "OUT",
            "Break Count",
            "Work Time",
            "Break Time",
            "Late (min)",
            "Early Exit (min)",
            "Manual Adjusted",
            "Day Closed",
        ]
        rows = []
        for s in sessions:
            emp = employees.get(s.employee_id)
            rows.append(
                [
                    emp.employee_code if emp else "",
                    emp.name if emp else "",
                    (emp.department if emp else "") or "",
                    s.status.value,
                    self._fmt(s.in_time),
                    self._fmt(s.break_out_time),
                    self._fmt(s.break_in_time),
                    self._fmt(s.out_time),
                    s.break_count,
                    format_seconds(s.total_work_seconds),
                    format_seconds(s.total_break_seconds),
                    s.late_minutes,
                    s.early_exit_minutes,
                    "Yes" if s.is_manually_adjusted else "No",
                    "Yes" if s.is_day_closed else "No",
                ]
            )
        wb = Workbook()
        write_sheet(wb, f"Daily {work_date.isoformat()}", headers, rows)
        return workbook_to_bytes(wb)

    # ------------------------------------------------------------------
    # Date-range aggregate — one row per employee, aggregated
    # ------------------------------------------------------------------

    def date_range_report(self, start: date, end: date) -> bytes:
        self._assert_range(start, end)
        sessions = self.daily_repo.list_in_range(start, end)

        by_employee: dict[int, list] = {}
        for s in sessions:
            by_employee.setdefault(s.employee_id, []).append(s)

        employees = self.employee_repo.get_by_ids(list(by_employee))
        total_days = (end - start).days + 1

        headers = [
            "Employee Code",
            "Name",
            "Department",
            "Days In Range",
            "Present Days",
            "Incomplete Days",
            "Absent Days",
            "Total Work Time",
            "Total Break Time",
            "Avg Work Time / Present Day",
            "Total Breaks",
            "Total Late (min)",
            "Total Early Exit (min)",
        ]
        rows = []
        for emp_id, slist in by_employee.items():
            emp = employees.get(emp_id)
            present = sum(1 for s in slist if s.status.value == "PRESENT")
            incomplete = sum(1 for s in slist if s.status.value == "INCOMPLETE")
            absent = total_days - present - incomplete
            work = sum(s.total_work_seconds for s in slist)
            brk = sum(s.total_break_seconds for s in slist)
            breaks_count = sum(s.break_count for s in slist)
            late = sum(s.late_minutes for s in slist)
            early = sum(s.early_exit_minutes for s in slist)
            avg = format_seconds(work // max(1, present)) if present else "0:00:00"
            rows.append(
                [
                    emp.employee_code if emp else "",
                    emp.name if emp else "",
                    (emp.department if emp else "") or "",
                    total_days,
                    present,
                    incomplete,
                    max(0, absent),
                    format_seconds(work),
                    format_seconds(brk),
                    avg,
                    breaks_count,
                    late,
                    early,
                ]
            )

        rows.sort(key=lambda r: (r[2] or "", r[0]))  # dept, then code
        wb = Workbook()
        write_sheet(wb, f"Range {start}_to_{end}", headers, rows)
        return workbook_to_bytes(wb)

    # ------------------------------------------------------------------
    # Monthly — convenience wrapper over date_range_report
    # ------------------------------------------------------------------

    def monthly_report(self, year: int, month: int) -> bytes:
        last_day = calendar.monthrange(year, month)[1]
        return self.date_range_report(date(year, month, 1), date(year, month, last_day))

    # ------------------------------------------------------------------
    # Employee-wise — one row per day in the range (ABSENT fill for gaps)
    # ------------------------------------------------------------------

    def employee_report(self, employee_id: int, start: date, end: date) -> bytes:
        self._assert_range(start, end)
        employee = self.employee_repo.get(employee_id)
        if employee is None:
            raise NotFoundError(f"Employee {employee_id} not found")
        sessions = self.daily_repo.list_for_employee_range(employee_id, start, end)
        existing = {s.work_date: s for s in sessions}

        headers = [
            "Date",
            "Status",
            "IN",
            "First Break Out",
            "First Break In",
            "OUT",
            "Break Count",
            "Work Time",
            "Break Time",
            "Late (min)",
            "Early Exit (min)",
            "Manual Adjusted",
        ]
        rows = []
        cur = start
        while cur <= end:
            s = existing.get(cur)
            if s is None:
                rows.append(
                    [cur.isoformat(), "ABSENT", "", "", "", "", 0, "0:00:00", "0:00:00", 0, 0, "No"]
                )
            else:
                rows.append(
                    [
                        cur.isoformat(),
                        s.status.value,
                        self._fmt(s.in_time),
                        self._fmt(s.break_out_time),
                        self._fmt(s.break_in_time),
                        self._fmt(s.out_time),
                        s.break_count,
                        format_seconds(s.total_work_seconds),
                        format_seconds(s.total_break_seconds),
                        s.late_minutes,
                        s.early_exit_minutes,
                        "Yes" if s.is_manually_adjusted else "No",
                    ]
                )
            cur += timedelta(days=1)

        wb = Workbook()
        write_sheet(wb, f"{employee.employee_code}", headers, rows)
        return workbook_to_bytes(wb)
