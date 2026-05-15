"""Excel report builders.

Reuses the existing ``services.logs.build_attendance_summaries`` /
``build_attendance_daily`` / ``build_attendance_range`` so the numbers in
the .xlsx file always match what the Reports page renders on screen — no
divergent SQL.
"""

from __future__ import annotations

import io
from datetime import date as date_cls
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import employees as employees_service
from .attendance import ShiftSettings
from .logs import (
    build_attendance_daily,
    build_attendance_range,
    build_attendance_summaries,
)

_HEADER_FILL = PatternFill("solid", fgColor="2F8F7B")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


# Column order common to all three sheet builders. Keeping it in one place so
# the daily / range / summary exports look identical to the operator.
_COLUMNS: list[tuple[str, str]] = [
    ("Date", "date"),
    ("Employee", "name"),
    ("Company", "company"),
    ("Status", "status"),
    ("Entry", "entry"),
    ("Exit", "exit"),
    ("Total Hours", "total_hours"),
    ("Total Working Hours", "total_working_hours"),
    ("Total Break Time", "total_break_time"),
    ("Late (min)", "late_minutes"),
    ("Early Exit (min)", "early_exit_minutes"),
]


def _safe_get(row: dict, key: str) -> object:
    if key == "company":
        # Daily/range rows from build_attendance_daily/range don't carry
        # company themselves — resolve it via the employees roster (same
        # path the existing summary endpoint uses).
        directory = row.get("__directory__")
        if directory is None:
            directory = employees_service.all_employees()
            row["__directory__"] = directory
        return employees_service.company_for(row.get("name") or "", employees=directory) or ""
    value = row.get(key)
    if value is None:
        return ""
    return value


def _apply_headers(sheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    sheet.freeze_panes = "A2"


def _autosize_columns(sheet) -> None:
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        # Cap at 40 chars wide to keep the sheet tidy on long names.
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def _write_rows(sheet, rows: list[dict], company_filter: Optional[str]) -> None:
    headers = [label for label, _ in _COLUMNS]
    _apply_headers(sheet, headers)
    target = (company_filter or "").strip().lower()
    out_row = 2
    for row in rows:
        if target:
            company = str(_safe_get(row, "company")).strip().lower()
            if company != target:
                continue
        for col_idx, (_label, key) in enumerate(_COLUMNS, start=1):
            sheet.cell(row=out_row, column=col_idx, value=_safe_get(row, key))
        out_row += 1
    _autosize_columns(sheet)


def build_daily_xlsx(
    *,
    target_date: date_cls,
    shift: ShiftSettings,
    company_filter: Optional[str],
) -> bytes:
    rows = build_attendance_daily(
        target_date=target_date,
        shift=shift,
        base_url="",
        expected_names=None,
    )
    wb = Workbook()
    sheet = wb.active
    sheet.title = f"Daily {target_date.isoformat()}"
    _write_rows(sheet, rows, company_filter)
    return _wb_to_bytes(wb)


def build_range_xlsx(
    *,
    start_date: date_cls,
    end_date: date_cls,
    shift: ShiftSettings,
    name_filter: Optional[str],
    company_filter: Optional[str],
) -> bytes:
    rows = build_attendance_range(
        start_date=start_date,
        end_date=end_date,
        shift=shift,
        base_url="",
        name_filter=name_filter,
    )
    wb = Workbook()
    sheet = wb.active
    title = f"Range {start_date.isoformat()} to {end_date.isoformat()}"
    sheet.title = title[:31]  # Excel sheet titles cap at 31 chars
    _write_rows(sheet, rows, company_filter)
    return _wb_to_bytes(wb)


def build_summary_xlsx(
    *,
    start_date: date_cls,
    end_date: date_cls,
    shift: ShiftSettings,
    name_filter: Optional[str],
    company_filter: Optional[str],
) -> bytes:
    rows = build_attendance_summaries(
        start_date=start_date,
        end_date=end_date,
        shift=shift,
        base_url="",
        name_filter=name_filter,
    )
    wb = Workbook()
    sheet = wb.active
    title = f"Attendance {start_date.isoformat()} to {end_date.isoformat()}"
    sheet.title = title[:31]
    _write_rows(sheet, rows, company_filter)
    return _wb_to_bytes(wb)


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
